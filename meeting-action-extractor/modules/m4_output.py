# -*- coding: utf-8 -*-
"""
M4 - 结果输出模块
职责：导出结构化表格，支持溯源跳转
输出格式：CSV、Excel
核心字段：content_type, description, owner, deadline, source_location, needs_review
"""

import csv
import io
import os
from typing import List, Optional

from modules.m2_extractor import ExtractionResult


# 核心输出字段定义（按赛道要求顺序）
OUTPUT_FIELDS = [
    ("content_type", "内容类型"),
    ("description", "决策/任务描述"),
    ("owner", "责任人"),
    ("deadline", "截止时间"),
    ("source_location", "原始字幕位置"),
    ("needs_review", "是否需人工确认"),
]

# 扩展字段（用于详细展示）
EXTENDED_FIELDS = [
    ("confidence", "置信度"),
    ("raw_text", "原始字幕文本"),
]


def results_to_dicts(results: List[ExtractionResult]) -> List[dict]:
    """将提取结果转为字典列表"""
    rows = []
    for r in results:
        row = {
            "content_type": r.content_type,
            "description": r.description,
            "owner": r.owner if r.owner else "(缺失)",
            "deadline": r.deadline if r.deadline else "(缺失)",
            "source_location": r.source_location,
            "needs_review": "是" if r.needs_review else "否",
            "confidence": f"{r.confidence:.0%}",
            "raw_text": r.raw_text,
        }
        rows.append(row)
    return rows


def export_csv(results: List[ExtractionResult], output_path: Optional[str] = None,
               include_extended: bool = False) -> str:
    """
    导出CSV格式结果
    输出：CSV文件路径或CSV字符串
    """
    rows = results_to_dicts(results)

    fields = OUTPUT_FIELDS[:]
    if include_extended:
        fields.extend(EXTENDED_FIELDS)

    headers = [f[1] for f in fields]
    keys = [f[0] for f in fields]

    if output_path:
        with open(output_path, 'w', newline='', encoding='utf-8-sig') as f:
            writer = csv.writer(f)
            writer.writerow(headers)
            for row in rows:
                writer.writerow([row.get(k, '') for k in keys])
        return output_path
    else:
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(headers)
        for row in rows:
            writer.writerow([row.get(k, '') for k in keys])
        return output.getvalue()


def export_excel(results: List[ExtractionResult], output_path: str) -> str:
    """导出Excel格式结果"""
    try:
        import pandas as pd
    except ImportError:
        # 降级为CSV
        csv_path = output_path.replace('.xlsx', '.csv')
        return export_csv(results, csv_path, include_extended=True)

    rows = results_to_dicts(results)
    fields = OUTPUT_FIELDS[:] + EXTENDED_FIELDS
    headers = [f[1] for f in fields]
    keys = [f[0] for f in fields]

    df = pd.DataFrame(rows)
    df = df[[k for k in keys]]
    df.columns = headers

    # 使用ExcelWriter设置格式
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='提取结果')

        # 获取工作表对象设置格式
        worksheet = writer.sheets['提取结果']

        # 设置列宽
        col_widths = {
            'A': 12,  # 内容类型
            'B': 40,  # 描述
            'C': 12,  # 责任人
            'D': 14,  # 截止时间
            'E': 25,  # 位置
            'F': 14,  # 需确认
            'G': 10,  # 置信度
            'H': 50,  # 原始文本
        }
        for col, width in col_widths.items():
            worksheet.column_dimensions[col].width = width

        # 标记需确认行
        from openpyxl.styles import PatternFill, Font, Alignment
        yellow_fill = PatternFill(start_color='FFFDE7', end_color='FFFDE7', fill_type='solid')
        for idx, row in enumerate(rows, start=2):  # 从第2行开始(跳过表头)
            if row.get('needs_review') == '是':
                for col_idx in range(1, len(fields) + 1):
                    cell = worksheet.cell(row=idx, column=col_idx)
                    cell.fill = yellow_fill

    return output_path


def get_summary(results: List[ExtractionResult]) -> dict:
    """生成统计概览"""
    total = len(results)
    decisions = sum(1 for r in results if r.content_type == "正式决策")
    actions = sum(1 for r in results if r.content_type == "行动项")
    discussions = sum(1 for r in results if r.content_type == "普通讨论")
    needs_review = sum(1 for r in results if r.needs_review)
    missing_owner = sum(1 for r in results if not r.owner and r.content_type in ("正式决策", "行动项"))
    missing_deadline = sum(1 for r in results if not r.deadline and r.content_type == "行动项")

    avg_confidence = sum(r.confidence for r in results) / total if total > 0 else 0

    return {
        "total": total,
        "decisions": decisions,
        "actions": actions,
        "discussions": discussions,
        "needs_review": needs_review,
        "missing_owner": missing_owner,
        "missing_deadline": missing_deadline,
        "avg_confidence": f"{avg_confidence:.0%}",
    }
